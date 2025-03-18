import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta

def get_monthly_costs(months_back=6, region='us-east-1'):
    try:
        client = boto3.client('ce', region_name=region)

        current_date = datetime.now()
        start_year = current_date.year
        start_month = current_date.month - months_back

        if start_month <= 0:
            start_year -= 1
            start_month = 12 + start_month

        start_date = f"{start_year}-{start_month:02d}-01"
        end_date = current_date.strftime("%Y-%m-%d")

        response = client.get_cost_and_usage(
            TimePeriod={'Start': start_date, 'End': end_date},
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=[{"Type": "DIMENSION", "Key": "SERVICE"}]  
        )

        monthly_costs = {}
        service_costs = {}
        total_cost = 0

        for result in response.get('ResultsByTime', []):
            start_date = result['TimePeriod']['Start']
            month_name = datetime.strptime(start_date, "%Y-%m-%d").strftime("%B %Y")

            monthly_costs[month_name] = 0
            service_costs[month_name] = {}

            for group in result.get('Groups', []):
                service = group['Keys'][0]
                cost = float(group.get('Metrics', {}).get('UnblendedCost', {}).get('Amount', 0))

                monthly_costs[month_name] += cost
                service_costs[month_name][service] = cost

            total_cost += monthly_costs[month_name]

        return {'monthly_costs': monthly_costs, 'service_costs': service_costs, 'total_cost': total_cost}

    except Exception as e:
        print(f"Error fetching AWS cost data: {e}")
        return {'monthly_costs': {}, 'service_costs': {}, 'total_cost': 0}

def generate_cost_comparison_report(cost_data, filename='aws_cost_comparison.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Cost Comparison"
    ws.append(["Month", "Cost ($)", "Change from Previous Month", "Change Percentage (%)"])

    monthly_costs = cost_data['monthly_costs']
    service_costs = cost_data['service_costs']
    sorted_months = sorted(monthly_costs.keys(), key=lambda x: datetime.strptime(x, "%B %Y"))
    prev_cost = None

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  

    for month in sorted_months:
        cost = round(monthly_costs[month], 2)
        change_text, change_percent_text = "N/A", "N/A"
        row_fill = None
        percent_fill = None

        if prev_cost is not None:
            cost_diff = round(cost - prev_cost, 2)
            percent_change = round((cost_diff / prev_cost) * 100, 2) if prev_cost != 0 else 0

            if cost_diff > 0:
                change_text = f"inc: ${cost_diff}"
                row_fill = yellow_fill
            else:
                change_text = f"dec: ${abs(cost_diff)}"
                row_fill = green_fill

            if percent_change > 100:
                percent_fill = red_fill
            elif percent_change >=50:
                percent_fill = orange_fill

            if percent_change > 0:
                change_percent_text = f"inc: {percent_change}%"
            else:
                change_percent_text = f"dec: {abs(percent_change)}%"

            # If cost increased by more than 50%, create a new sheet with service details
            if percent_change >=50:
                service_ws = wb.create_sheet(title=f"{month} Breakdown")
                service_ws.append(["Service", "Cost ($)"])

                total_service_cost = 0  
                for service, service_cost in service_costs[month].items():
                    service_ws.append([service, round(service_cost, 2)])
                    total_service_cost += service_cost  

                # Append total cost row in service breakdown sheet
                service_ws.append(["Total", round(total_service_cost, 2)])  
                for col in range(1, 3):
                    service_ws.cell(row=service_ws.max_row, column=col).font = Font(bold=True)

        ws.append([month, cost, change_text, change_percent_text])

        if row_fill:
            ws.cell(row=ws.max_row, column=3).fill = row_fill  
        if percent_fill:
            ws.cell(row=ws.max_row, column=4).fill = percent_fill  

        prev_cost = cost

    total_cost = round(cost_data['total_cost'], 2)
    ws.append(["Total", total_cost, "", ""])  

    for col in range(1, 5):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    wb.save(filename)

def main():
    cost_data = get_monthly_costs()
    filename = f"aws_cost_comparison_{datetime.now().strftime('%Y%m%d')}.xlsx"
    generate_cost_comparison_report(cost_data, filename)
    print(f"Report generated: {filename}")

if __name__ == "__main__":
    main()
