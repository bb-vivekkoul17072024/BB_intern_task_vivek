import re
from datetime import datetime, timezone
import boto3
from openpyxl import load_workbook
input_workbook = load_workbook("missing_tags_instances.xlsx")
input_sheet = input_workbook.active
input_data = {}
instance_id_pattern = re.compile(r"^i-[a-f0-9]{8,17}$")  # Regex to validate instance ID format
for row in input_sheet.iter_rows(min_row=2, values_only=True):
    instance_id = row[0]
    if instance_id and instance_id_pattern.match(instance_id):  # Validate instance ID format
        input_data[instance_id] = {
            'InstanceType': row[1] if row[1] is not None else "",
            'Name': row[2] if row[2] is not None else "",
            'application_module': row[3] if row[3] is not None else "",
            'team': row[4] if row[4] is not None else "",
            'patch': row[5] if row[5] is not None else "",
            'os': row[6] if row[6] is not None else "",
            'ssm': row[7] if row[7] is not None else "",
            'State': row[8] if row[8] is not None else "",
        }
    elif instance_id:
        print(f"Invalid instance ID format: {instance_id}")
# print(input_data)--working
regions = {"ap-south-1", "us-east-1"}
for region in regions:
    ec2 = boto3.client('ec2', region_name=region)
    response = ec2.describe_instances()
    for res in response['Reservations']:
        for instance in res['Instances']:
            instance_id = instance['InstanceId']
            instance_state = instance['State']['Name']
            if instance_id in input_data and instance_state == 'running':
                instance_tags_dict = {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])}
                required_tags = input_data[instance_id]
                valid_tag_keys = {'Name', 'application_module', 'team', 'patch', 'os', 'ssm'}
                missing_or_empty_tags = {
                    key: value for key, value in required_tags.items()
                    if key in valid_tag_keys and (key not in instance_tags_dict or instance_tags_dict[key] == "") and value != ""
                }
                if missing_or_empty_tags:
                    try:
                        ec2.create_tags(
                            Resources=[instance_id],
                            Tags=[{'Key': key, 'Value': value} for key, value in missing_or_empty_tags.items()]
                        )
                        print(f"Updated tags for instance {instance_id}: {missing_or_empty_tags}")
                    except Exception as e:
                        print(f"Error updating tags for instance {instance_id}: {e}")