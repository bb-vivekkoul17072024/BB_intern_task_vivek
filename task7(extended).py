import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def get_linked_accounts():
    client = boto3.client("organizations")
    accounts = []
    paginator = client.get_paginator("list_accounts")
    
    for page in paginator.paginate():
        for account in page["Accounts"]:
            if account["Status"] == "ACTIVE":
                accounts.append(account["Id"])
    
    return accounts

def get_monthly_costs(account_id, months_back=6, region="us-east-1"):
    client = boto3.client("ce", region_name=region)
    
    current_date = datetime.now()
    start_year = current_date.year
    start_month = current_date.month - months_back
    
    if start_month <= 0:
        start_year -= 1
        start_month = 12 + start_month
    
    start_date = f"{start_year}-{start_month:02d}-01"
    end_date = current_date.strftime("%Y-%m-%d")

    response = client.get_cost_and_usage(
        TimePeriod={"Start": start_date, "End": end_date},
        Granularity="MONTHLY",
        Metrics=["UnblendedCost"],
        GroupBy=[{"Type": "DIMENSION", "Key": "SERVICE"}],
        Filter={"Dimensions": {"Key": "LINKED_ACCOUNT", "Values": [account_id]}},
    )
    
    monthly_costs = {}
    service_costs = {}
    total_cost = 0

    for result in response.get("ResultsByTime", []):
        month_name = datetime.strptime(result["TimePeriod"]["Start"], "%Y-%m-%d").strftime("%B %Y")
        monthly_costs[month_name] = 0
        service_costs[month_name] = {}

        for group in result.get("Groups", []):
            service = group["Keys"][0]
            cost = float(group.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", 0))
            
            monthly_costs[month_name] += cost
            service_costs[month_name][service] = cost

        total_cost += monthly_costs[month_name]
    
    return {"monthly_costs": monthly_costs, "service_costs": service_costs, "total_cost": total_cost}

def generate_multi_account_report(account_costs, filename="aws_multi_account_costs.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multi-Account Cost Report"

    ws.append(["Account ID", "Month", "Cost ($)", "Change from Previous Month", "Change Percentage (%)"])
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for account_id, cost_data in account_costs.items():
        monthly_costs = cost_data["monthly_costs"]
        service_costs = cost_data["service_costs"]
        sorted_months = sorted(monthly_costs.keys(), key=lambda x: datetime.strptime(x, "%B %Y"))
        prev_cost = None

        for month in sorted_months:
            cost = round(monthly_costs[month], 2)
            change_text, change_percent_text = "N/A", "N/A"
            row_fill = None
            percent_fill = None
            percent_change = 0

            if prev_cost is not None:
                cost_diff = round(cost - prev_cost, 2)
                percent_change = round((cost_diff / prev_cost) * 100, 2) if prev_cost != 0 else 0
                
                if cost_diff > 0:
                    change_text = f"inc: ${cost_diff}"
                    row_fill = yellow_fill
                else:
                    change_text = f"dec: ${abs(cost_diff)}"
                    row_fill = green_fill

                if percent_change > 100:
                    percent_fill = red_fill
                elif percent_change >= 50:
                    percent_fill = orange_fill
                
                change_percent_text = f"{percent_change}%"

            ws.append([account_id, month, cost, change_text, change_percent_text])
            row_index = ws.max_row
            if row_fill:
                for col in range(2, 6):
                    ws.cell(row=row_index, column=col).fill = row_fill
            if percent_fill:
                ws.cell(row=row_index, column=5).fill = percent_fill

            prev_cost = cost

            if percent_change >= 50:
                service_ws = wb.create_sheet(title=f"{account_id}_{month}")
                service_ws.append(["Service", "Cost ($)"])
                total_service_cost = 0
                
                for service, service_cost in service_costs[month].items():
                    service_ws.append([service, round(service_cost, 2)])
                    total_service_cost += round(service_cost, 2)
                
                # Add total cost at the end of the sheet
                service_ws.append(["Total Cost", total_service_cost])
                bold_font = Font(bold=True)
                service_ws.cell(row=service_ws.max_row, column=1).font = bold_font
                service_ws.cell(row=service_ws.max_row, column=2).font = bold_font

    wb.save(filename)

def main():
    accounts = get_linked_accounts()
    account_costs = {}

    for account_id in accounts:
        print(f"Fetching cost data for account: {account_id}...")
        account_costs[account_id] = get_monthly_costs(account_id)

    filename = f"aws_multi_account_costs_{datetime.now().strftime('%Y%m%d')}.xlsx"
    generate_multi_account_report(account_costs, filename)
    print(f"Multi-account cost report generated: {filename}")

if __name__ == "__main__":
    main()
